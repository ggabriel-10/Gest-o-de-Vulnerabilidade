import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from tkinter import Tk, filedialog, messagebox

# Configurar a janela do Tkinter para selecionar múltiplos arquivos de entrada
root = Tk()
root.withdraw()
input_files = filedialog.askopenfilenames(title="Selecione os arquivos de entrada", filetypes=[("CSV e Excel", "*.csv *.xlsx")])

# Verificar se algum arquivo foi selecionado
if not input_files:
    print("Nenhum arquivo selecionado. Encerrando...")
    exit()

# Iterar sobre os arquivos selecionados para definir um nome de saída para cada um
for input_file in input_files:
    # Exibir a caixa de diálogo informando o usuário sobre o arquivo sendo editado
    messagebox.showinfo("Edição de arquivos", f"Você está editando o arquivo: {input_file}")

    # Solicitar o nome do arquivo de saída
    output_file = filedialog.asksaveasfilename(title=f"Salvar {input_file}", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
    
    if not output_file:
        print(f"Nenhum arquivo de saída definido para {input_file}. Continuando para o próximo arquivo...")
        continue

    # URL da imagem
    image_url = 'https://i.ibb.co/5sdXx72/SEK-RGB-Horizontal-Descritivo-Negativo.png'

    # Baixar a imagem
    response = requests.get(image_url)
    img_data = BytesIO(response.content)

    # Ler o arquivo, identificando se é CSV ou Excel
    if input_file.endswith('.csv'):
        df = pd.read_csv(input_file, skiprows=7, low_memory=False)  # Evita DtypeWarning
    else:
        df = pd.read_excel(input_file, skiprows=7)

    # Filtrar apenas as linhas onde a coluna 'Type' é igual a 'Vuln'
    df = df[df['Type'].str.contains('Vuln', case=False, na=False)]

    # Definir as colunas desejadas na ordem especificada
    desired_columns = [
        'Title', 'Severity', 'IP', 'DNS', 'NetBIOS', 'OS', 
        'Port', 'Protocol', 'Results', 'Threat', 'Impact', 
        'Solution', 'CVE ID'
    ]

    # Filtrar o DataFrame para manter apenas as colunas desejadas
    df = df[desired_columns]

    # Trocar o nome das colunas
    df.rename(columns={'CVE ID': 'CVE Numbers', 'IP': 'Host IP'}, inplace=True)

    # Substituir valores na coluna Severity
    severity_mapping = {
        1: 'Minimal',
        2: 'Medium',
        3: 'Serious',
        4: 'Critical',
        5: 'Urgent'
    }
    df['Severity'] = df['Severity'].replace(severity_mapping)

    # Excluir linhas onde a coluna 'Severity' está em branco
    df = df[df['Severity'].notna()]  # Remove todas as linhas onde 'Severity' é NaN (vazio)

    # Ordenar a coluna Severity
    severity_order = ['Urgent', 'Critical', 'Serious', 'Medium', 'Minimal']
    df['Severity'] = pd.Categorical(df['Severity'], categories=severity_order, ordered=True)
    df = df.sort_values('Severity')

    # Salvar o DataFrame formatado em um novo arquivo Excel
    df.to_excel(output_file, index=False)

    # Abrir o arquivo Excel formatado
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # 1. Adicionar uma linha acima da primeira linha atual
    ws.insert_rows(1)

    # 2. Limpar e formatar a linha de título
    ws["A1"].value = ""  
    ws["B1"].value = "PROOF SOC / Gestão de Vulnerabilidades - Controle"
    ws["A1"].fill = PatternFill(start_color='1F1F1F', fill_type='solid')
    ws["B1"].fill = PatternFill(start_color='1F1F1F', fill_type='solid')
    ws["A1"].font = Font(color='CCCCCC')
    ws["B1"].font = Font(name='Segoe UI', size=16, color='CCCCCC', bold=False)  # Removido negrito
    ws.row_dimensions[1].height = 79.50  # Altura da linha 1

    # 3. Mesclar e alinhar células de B1 a M1
    ws.merge_cells('B1:M1')  # Mesclar de B a M
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')  # Alinhamento à esquerda

    # 4. Adicionar a imagem na célula A1
    img = Image(img_data)
    img.height = 0.85 * 72
    img.width = 2.45 * 72
    img.anchor = 'A1'
    ws.add_image(img)

    # Estilizar a linha de título (B1 a M1)
    for col in range(2, 14):  # Colunas B (2) até M (13)
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='1F1F1F', fill_type='solid')  # Fundo preto
        cell.font = Font(color='CCCCCC', name='Segoe UI', size=16, bold=False)  # Removido negrito
        cell.alignment = Alignment(horizontal='left', vertical='center')  # Alinhamento à esquerda

    # Definindo a altura da linha 2
    ws.row_dimensions[2].height = 60.00  # Altura da linha 2

    # Estilização da linha 2
    for cell in ws["2:2"]:  # Altera para a linha de cabeçalho correta
        cell.font = Font(color="FFFFFF", name="Segoe UI", size=11)  # Cor da fonte branca e fonte Segoe UI
        cell.fill = PatternFill(start_color="262626", fill_type="solid")  # Fundo com a cor #262626
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinhar ao centro

    # Aplicando o preenchimento alternado de cinza e branco a partir da linha 3, exceto para a coluna B (Severity)
    for row in range(3, ws.max_row + 1):
        fill_color = "F2F2F2" if row % 2 == 0 else "FFFFFF"  # Cor cinza claro e branco alternados
        for col in range(1, len(desired_columns) + 1):
            if col != 2:  # Ignorar a coluna Severity
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, fill_type="solid")

    # Definindo cores para a coluna Severity
    severity_colors = {
        'Urgent': 'FF0000',  # Vermelho
        'Critical': 'FFA500',  # Laranja
        'Medium': '9966FF',  # Lilás
        'Minimal': 'A6A6A6',  # Cinza
        'Serious': 'FFFF00'  # Amarelo
    }

    # Aplicando cores e formatando
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=len(desired_columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Alinhar ao centro e habilitar quebra de texto
            cell.font = Font(name="Segoe UI", size=11)  # Definindo fonte

            if cell.column == 2:  # Coluna Severity
                severity_value = cell.value
                if severity_value in severity_colors:
                    cell.fill = PatternFill(start_color=severity_colors[severity_value], end_color=severity_colors[severity_value], fill_type='solid')

    # Definindo a altura das linhas de 3 para baixo
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 409.50  # Altura das linhas a partir da linha 3

    # Definindo a largura das colunas
    column_widths = {
    1: 39.27,  # Title
    2: 15.18,  # Severity
    3: 24.82,  # Host IP
    4: 15.18,  # DNS
    5: 15.18,  # NetBIOS
    6: 15.18,  # OS
    7: 12.64,  # Port
    8: 14.91,  # Protocol
    9: 14.91,  # Results
    10: 36.18,  # Threat
    11: 36.18,  # Impact
    12: 78.82,  # Solution
    13: 24.82   # CVE Numbers
    }

    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Adicionar bordas a todas as células de A2 a M
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.border = border

    # Salvar o arquivo com bordas aplicadas
    wb.save(output_file)

    # Exibir mensagem informando que a edição foi concluída
    messagebox.showinfo("Edição concluída", f"O arquivo foi salvo como: {output_file}")
