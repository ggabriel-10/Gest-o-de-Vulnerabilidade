import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
import requests  # Biblioteca para baixar a imagem a partir do link
from io import BytesIO

# Definir caminhos de entrada e saída
input_excel = 'teste.xlsx'
output_excel = 'formatted_qualys.xlsx'
image_url = 'https://i.ibb.co/5sdXx72/SEK-RGB-Horizontal-Descritivo-Negativo.png'

# 1. Ler o arquivo Excel e identificar a linha "RESULTS"
df = pd.read_excel(input_excel, header=None)  # Ler sem cabeçalho para capturar todas as linhas
results_index = df[df.apply(lambda row: row.astype(str).str.contains("RESULTS", case=False).any(), axis=1)].index[0]

# 2. Remover todas as linhas acima da linha que contém "RESULTS"
df = df.iloc[results_index + 1:].reset_index(drop=True)

# 3. Definir os nomes das colunas
df.columns = df.iloc[0]  # Usa a primeira linha como cabeçalho
df = df[1:]  # Remove a linha do cabeçalho

# 4. Filtrar as colunas de interesse
columns_to_keep = [
    'VULNERABILITY', 'Severity Level', 'Description', 'Impact', 'Solution',
    'Url', 'Access Path', 'Response #1', 'OWASP', 'CWE', 'Title'  # Inclui 'Title' temporariamente
]
df = df[columns_to_keep]

# 5. Renomear colunas conforme especificado
df = df.rename(columns={
    'VULNERABILITY': 'Vulnerabilidade',
    'Severity Level': 'Severidade',
    'Description': 'Descrição',
    'Impact': 'Impacto',
    'Solution': 'Mitigação',
    'Url': 'URL',
    'Access Path': 'Caminho',
    'Response #1': 'Resposta',
    'OWASP': 'OWASP',
    'CWE': 'CWE',
    'Title': 'Title'  # Renomeie se necessário
})

# 6. Manter apenas as linhas que possuem "Vulnerabilidade"
df = df[df['Vulnerabilidade'].str.contains("VULNERABILITY", na=False)]

# Copiar os valores da coluna 'Title' para 'Vulnerabilidade'
df['Vulnerabilidade'] = df['Title']

# 7. Remover a coluna 'Title' do DataFrame
df = df.drop(columns=['Title'])

# 8. Ajustar valores da coluna 'Severidade'
severity_mapping = {
    1: 'Minimal',
    2: 'Medium',
    3: 'Serious',
    4: 'Critical',
    5: 'Urgent'
}
df['Severidade'] = df['Severidade'].replace(severity_mapping)

# 9. Ordenar a coluna "Severidade" na ordem desejada
severity_order = ['Urgent', 'Critical', 'Serious', 'Medium', 'Minimal']
df['Severidade'] = pd.Categorical(df['Severidade'], categories=severity_order, ordered=True)
df = df.sort_values('Severidade')

# 10. Salvar em um novo arquivo Excel
df.to_excel(output_excel, index=False)

# 11. Reabrir o arquivo para aplicar estilização com OpenPyxl
wb = openpyxl.load_workbook(output_excel)
ws = wb.active

# 12. Adicionar uma linha acima da primeira linha atual
ws.insert_rows(1)

# 13. Limpar e formatar a linha de título
ws["A1"].value = ""  
ws["B1"].value = "SEK SOC / Gestão de Vulnerabilidades - Controle"
ws["A1"].fill = PatternFill(start_color='1F1F1F', fill_type='solid')
ws["B1"].fill = PatternFill(start_color='1F1F1F', fill_type='solid')
ws["A1"].font = Font(color='CCCCCC')
ws["B1"].font = Font(name='Segoe UI', size=16, color='CCCCCC', bold=True)
ws.row_dimensions[1].height = 79.50
ws.row_dimensions[2].height = 60.00

# 14. Ajustar estilos de células
default_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 14.5. Aplicar "Unir e Centrar" nas colunas B a J
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=10)  # B1:J1
ws['B1'].alignment = Alignment(horizontal='center', vertical='center')  # Centralizar

# 14.6. Alinhar a esquerda as células da linha 1 nas colunas B a J
for col in range(2, 11):  # Colunas B a J
    ws.cell(row=1, column=col).alignment = Alignment(horizontal='left', vertical='center')

# 15. Ajustar alinhamento da coluna A até J, linha 2 em diante (Centralizado)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):  # Colunas A até J (1 até 10)
    for cell in row:
        if isinstance(cell.value, str):
            cell.value = cell.value.replace("\n", " ")  # Removendo quebras de linha sem inserir novas
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)  # Centralizado
        cell.font = Font(name='Calibri', size=11)

# 16. Estilizar cabeçalho
for cell in ws[2]:
    cell.font = Font(color="CCCCCC", bold=True, name='Segoe UI', size=11)
    cell.fill = PatternFill(start_color="1F1F1F", fill_type="solid")

# 16.1. Aplicar a fonte Calibri, tamanho 11, e a cor #FFFFFF para as colunas A a J na linha 2
for col in range(1, 11):  # Colunas A (1) a J (10)
    ws.cell(row=2, column=col).font = Font(name='Calibri', size=11, color='FFFFFF')  # Cor branca

# 17. Aplicar cores e altura das linhas com base na severidade
severity_colors = {
    'Minimal': 'D3D3D3',
    'Medium': '7C4DFF',
    'Serious': 'FFFF00',
    'Critical': 'FFA500',
    'Urgent': 'FF0000'
}

for row in ws.iter_rows(min_row=3, min_col=2, max_col=2):
    for cell in row:
        severity_value = cell.value
        if severity_value in severity_colors:
            cell.fill = PatternFill(start_color=severity_colors[severity_value], fill_type='solid')
        ws.row_dimensions[cell.row].height = 87

# 18. Aplicar cor #7C4DFF para células que contêm "Medium" e mudar a fonte para preto
medium_fill = PatternFill(start_color='7C4DFF', fill_type='solid')  
medium_font = Font(color='000000')

for row in ws.iter_rows(min_row=3, min_col=2, max_col=2):  
    for cell in row:
        if cell.value == 'Medium':
            cell.fill = medium_fill
            cell.font = medium_font

# 19. Definir larguras das colunas A, B e C
ws.column_dimensions['A'].width = 84.82
ws.column_dimensions['B'].width = 19.91
ws.column_dimensions['C'].width = 79.00

# 20. Ajustar o formato das colunas A, B e C
for cell in ws['A']:
    cell.number_format = '#,##0.00'
for cell in ws['B']:
    cell.number_format = '#,##0.00'
for cell in ws['C']:
    cell.number_format = '#,##0.00'

# 21. Baixar e inserir a imagem a partir do link
response = requests.get(image_url)
img_data = BytesIO(response.content)
img = Image(img_data)

# Ajustar tamanho da imagem
img.height = 0.85 * 72
img.width = 2.45 * 72
img.anchor = 'A1'
ws.add_image(img)

# 22. Definir larguras e formatos das colunas D, E, F, G, H, I, J
ws.column_dimensions['D'].width = 79.00  # 876 pixels
ws.column_dimensions['E'].width = 43.82  # 489 pixels
ws.column_dimensions['F'].width = 33.91  # 380 pixels
ws.column_dimensions['G'].width = 38.82  # 434 pixels
ws.column_dimensions['H'].width = 88.36  # 979 pixels
ws.column_dimensions['I'].width = 46.36  # 517 pixels
ws.column_dimensions['J'].width = 20.18  # 229 pixels

# 23. Ajustar o formato das colunas D, E, F, G, H, I, J
for cell in ws['D']:
    cell.number_format = '#,##0.00'
for cell in ws['E']:
    cell.number_format = '#,##0.00'
for cell in ws['F']:
    cell.number_format = '#,##0.00'
for cell in ws['G']:
    cell.number_format = '#,##0.00'
for cell in ws['H']:
    cell.number_format = '#,##0.00'
for cell in ws['I']:
    cell.number_format = '#,##0.00'
for cell in ws['J']:
    cell.number_format = '#,##0.00'

# 24. Criar tabela com estilo "Médio, Branco, estilo de tabela intermédio 15"
table_ref = f"A2:{get_column_letter(len(columns_to_keep) - 1)}{ws.max_row}"
table = Table(displayName="QualysTable", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

# 25. Salvar arquivo com formatação aplicada
wb.save(output_excel)
print(f"Arquivo formatado salvo como '{output_excel}' com sucesso!")
