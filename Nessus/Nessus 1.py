import csv
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilenames, asksaveasfilename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd

def reformat_csv(input_file, output_file):
    desired_columns = [
        'Name', 'Severity', 'VPR Score', 'CVSS v4.0 Base Score', 'CVSS v3.0 Base Score', 'CVSS v2.0 Base Score', 'Host', 'Port', 'Protocol',
        'Synopsis', 'Description', 'Solution', 'Plugin Output'
    ]

    temp_csv = "temp_output.csv"  # Arquivo temporário para salvar o CSV formatado

    with open(input_file, mode='r', newline='', encoding='utf-8') as infile:
        reader = csv.DictReader(infile)

        with open(temp_csv, mode='w', newline='', encoding='utf-8') as outfile:
            writer = csv.DictWriter(outfile, fieldnames=desired_columns)
            writer.writeheader()

            for row in reader:
                try:
                    # Limpeza dos campos
                    for col in ['VPR Score', 'CVSS v4.0 Base Score', 'CVSS v3.0 Base Score', 'CVSS v2.0 Base Score']:
                        if not row[col] or not row[col].strip():
                            row[col] = ''

                    # Ignorar linhas com todas as pontuações vazias
                    if not row['VPR Score'] and not row['CVSS v4.0 Base Score'] and not row['CVSS v3.0 Base Score'] and not row['CVSS v2.0 Base Score']:
                        continue

                    # Determinar Severity com base nos valores
                    if row['VPR Score']:
                        vpr_score = float(row['VPR Score'])
                        if 0.1 <= vpr_score <= 3.9:
                            row['Severity'] = 'Low'
                        elif 4.0 <= vpr_score <= 6.9:
                            row['Severity'] = 'Medium'
                        elif 7.0 <= vpr_score <= 8.9:
                            row['Severity'] = 'High'
                        elif 9.0 <= vpr_score <= 10.0:
                            row['Severity'] = 'Critical'
                        else:
                            row['Severity'] = ''
                    else:
                        for score_col in ['CVSS v4.0 Base Score', 'CVSS v3.0 Base Score', 'CVSS v2.0 Base Score']:
                            try:
                                if row[score_col]:
                                    cvss_score = float(row[score_col])
                                    if 0.1 <= cvss_score <= 3.9:
                                        row['Severity'] = 'Low'
                                    elif 4.0 <= cvss_score <= 6.9:
                                        row['Severity'] = 'Medium'
                                    elif 7.0 <= cvss_score <= 8.9:
                                        row['Severity'] = 'High'
                                    elif 9.0 <= cvss_score <= 10.0:
                                        row['Severity'] = 'Critical'
                                    else:
                                        row['Severity'] = ''
                                    break
                            except ValueError:
                                pass

                    if not row['Severity']:
                        continue

                    # Filtrar as colunas desejadas
                    filtered_row = {col: row[col] for col in desired_columns if col in row}
                    writer.writerow(filtered_row)

                except KeyError as e:
                    print(f"Coluna ausente ignorada: {e}")
                except ValueError as e:
                    print(f"Erro ao processar linha: {e}")

    # Convertendo o CSV para DataFrame para manipulação
    df = pd.read_csv(temp_csv)

    # Definindo a ordem da severidade
    severity_order = ['Critical', 'High', 'Medium', 'Low'] 
    df['Severity'] = pd.Categorical(df['Severity'], categories=severity_order, ordered=True)
    df = df.sort_values('Severity')

    # Salvando o DataFrame de volta no arquivo CSV temporário
    df.to_csv(temp_csv, index=False)

    # Convertendo o CSV para Excel e aplicando formatação
    wb = Workbook()
    ws = wb.active

    with open(temp_csv, mode='r', newline='', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        for row in reader:
            ws.append(row)

    # URL da imagem
    image_url = 'https://i.ibb.co/5sdXx72/SEK-RGB-Horizontal-Descritivo-Negativo.png'

    # Baixar a imagem
    response = requests.get(image_url)
    img_data = BytesIO(response.content)

    # Adicionar uma linha acima da primeira linha atual
    ws.insert_rows(1)

    # Limpar e formatar a linha de título
    ws["A1"].value = ""  
    ws["B1"].value = "SEK SOC / Gestão de Vulnerabilidades - Controle"
    ws["A1"].fill = PatternFill(start_color='1F1F1F', fill_type='solid')
    ws["B1"].fill = PatternFill(start_color='1F1F1F', fill_type='solid')
    ws["A1"].font = Font(color='CCCCCC')
    ws["B1"].font = Font(name='Segoe UI', size=16, color='CCCCCC', bold=False)  # Removido negrito
    ws.row_dimensions[1].height = 79.50  # Altura da linha 1

    # Mesclar e alinhar células de B1 a M1
    ws.merge_cells('B1:M1')  # Mesclar de B a M
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')  # Alinhamento à esquerda

    # Adicionar a imagem na célula A1
    img = Image(img_data)
    img.height = 0.85 * 72
    img.width = 2.45 * 72
    img.anchor = 'A1'
    ws.add_image(img)

    # Estilizar a linha de título (B1 a M1)
    for col in range(2, 14):  # Colunas B (2) até M (13)
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='1F1F1F', fill_type='solid')  # Fundo preto
        cell.font = Font(color='CCCCCC', name='Segoe UI', size=16, bold=False)  # Removido negrito
        cell.alignment = Alignment(horizontal='left', vertical='center')  # Alinhamento à esquerda

    # Alterando o nome da coluna 'Name' para 'Title' na linha 2
    ws["A2"].value = "Title"

    # Definindo a altura da linha 2
    ws.row_dimensions[2].height = 60.00  # Altura da linha 2

    # Estilização da linha 2
    for cell in ws["2:2"]:  # Altera para a linha de cabeçalho correta
        cell.font = Font(color="FFFFFF", name="Segoe UI", size=11)  # Cor da fonte branca e fonte Segoe UI
        cell.fill = PatternFill(start_color="262626", fill_type="solid")  # Fundo com a cor #262626
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinhar ao centro

    # Aplicando o preenchimento alternado de cinza e branco a partir da linha 3
    for row in range(3, ws.max_row + 1):
        fill_color = "F2F2F2" if row % 2 == 0 else "FFFFFF"  # Cor cinza claro e branco alternados
        for col in range(1, len(desired_columns) + 1):
            if col != 2:  # Ignorar a coluna Severity
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, fill_type="solid")

    # Definindo cores para a coluna Severity
    severity_colors = {
        'Critical': 'FF0000',  # Vermelho
        'High': 'FFA500',  # Laranja
        'Medium': '9966FF',  # Lilás
        'Low': 'A6A6A6'  # Cinza
    }

    # Aplicando cores e formatando com base na nova ordem
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=len(desired_columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Alinhar ao centro e habilitar quebra de texto
            cell.font = Font(name="Segoe UI", size=11)  # Definindo fonte

            if cell.column == 2:  # Coluna Severity
                severity_value = cell.value
                if severity_value in severity_colors:
                    # Alterar a cor de fundo com base no valor de severidade
                    cell.fill = PatternFill(start_color=severity_colors[severity_value], end_color=severity_colors[severity_value], fill_type='solid')

    # Definindo a altura das linhas de 3 para baixo
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 409.50  # Altura das linhas a partir da linha 3

    # Definindo a largura das colunas
    column_widths = {
        'A': 39.27,  # Name
        'B': 15.18,  # Severity
        'C': 14.55,  # VPR Score
        'D': 25.18,  # CVSS v4.0 Base Score
        'E': 25.18,  # CVSS v3.0 Base Score
        'F': 25.18,  # CVSS v2.0 Base Score
        'G': 17.73,  # Host
        'H': 17.73,  # Port
        'I': 17.73,  # Protocol
        'J': 41.91,  # Synopsis
        'K': 41.91,  # Description
        'L': 41.91,  # Solution
        'M': 20.00   # Plugin Output
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Aplicando a formatação de tabela
    table_range = f"A2:M{ws.max_row}"
    table = Table(displayName="Table1", ref=table_range)

    # Estilo da Tabela
    style = TableStyleInfo(
        name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Salvando o arquivo Excel
    wb.save(output_file)

# Interface gráfica para selecionar arquivos
def select_file():
    Tk().withdraw()  # Evitar a janela principal do Tkinter
    input_files = askopenfilenames(title="Escolha os arquivos CSV de entrada", filetypes=[("CSV Files", "*.csv")])
    if not input_files:
        return None, None

    # Pedir para o usuário selecionar o nome de saída para cada arquivo
    output_files = []
    for input_file in input_files:
        output_file = asksaveasfilename(
            title="Escolha o arquivo de saída", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=input_file.split('/')[-1].replace('.csv', '.xlsx')
        )

        # Perguntar se o usuário deseja continuar com o arquivo
        root = Tk()
        root.withdraw()  # Ocultar janela principal
        if messagebox.askyesno("Confirmar", f"Deseja continuar com o arquivo selecionado {input_file}?"):
            output_files.append(output_file)

    return input_files, output_files

# Processo de conversão
def main():
    input_files, output_files = select_file()
    if input_files and output_files:
        for input_file, output_file in zip(input_files, output_files):
            reformat_csv(input_file, output_file)

if __name__ == "__main__":
    main()
